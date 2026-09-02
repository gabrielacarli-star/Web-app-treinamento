#!/usr/bin/env python3
"""
Consolida todos os bancos JSON da operacao numa planilha unica.

Uso:
    python3 consolidar_planilha.py [pasta_raiz] [saida.xlsx]

Padrao: pasta_raiz = "." e saida = "banco-criativo.xlsx"

Le, quando existirem:
    banco/criativos.json          dissecacao-de-criativo   (1 linha por ad dissecado)
    banco/virais.json             mineracao-formatos-virais (organico TT/YT/IG)
    banco/publico.json            estudo-de-publico
    mineracao/angulos.json        mineracao-angulo
    mineracao/clickbait.json      mineracao-clickbait
    mineracao/formatos.json       mineracao-formatos-virais
    mineracao/benchmark.json      benchmarking-mercado
    banco-de-fatias.json          o ativo persistente

Gera abas: Resumo, Criativos, Angulos, Avatares, Formatos, Hooks, Publico,
Virais, Anunciantes, Fatias.

Nao inventa dado. Campo ausente vira vazio. Arquivo ausente vira aba vazia
com o cabecalho, pra planilha nunca quebrar por falta de fonte.
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FONTE = "Arial"
CAB_FILL = PatternFill("solid", fgColor="1F2937")
CAB_FONT = Font(name=FONTE, bold=True, color="FFFFFF", size=10)
CEL_FONT = Font(name=FONTE, size=10)
LINK_FONT = Font(name=FONTE, size=10, color="0563C1", underline="single")


def carregar(caminho: Path):
    """Le um JSON. Devolve [] se nao existir ou estiver quebrado."""
    if not caminho.exists():
        return []
    try:
        with open(caminho, encoding="utf-8") as fh:
            return json.load(fh)
    except (json.JSONDecodeError, OSError) as erro:
        print(f"  aviso: {caminho} ilegivel ({erro}) - tratado como vazio")
        return []


def pegar(obj, *caminho, padrao=""):
    """Acessa chave aninhada sem estourar. pegar(o,'fonte','link_funil')."""
    atual = obj
    for chave in caminho:
        if not isinstance(atual, dict):
            return padrao
        atual = atual.get(chave)
        if atual is None:
            return padrao
    if isinstance(atual, list):
        return ", ".join(str(x) for x in atual)
    return atual


def escrever_aba(wb, titulo, colunas, linhas, cols_link=()):
    ws = wb.create_sheet(titulo)
    ws.append(colunas)
    for celula in ws[1]:
        celula.fill = CAB_FILL
        celula.font = CAB_FONT
        celula.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22

    for linha in linhas:
        ws.append(linha)

    idx_link = {colunas.index(c) + 1 for c in cols_link if c in colunas}
    for row in ws.iter_rows(min_row=2):
        for celula in row:
            celula.font = CEL_FONT
            celula.alignment = Alignment(vertical="top", wrap_text=True)
            if celula.column in idx_link and isinstance(celula.value, str) \
                    and celula.value.startswith("http"):
                celula.hyperlink = celula.value
                celula.font = LINK_FONT

    for i, nome in enumerate(colunas, start=1):
        largura = 42 if nome.startswith("link") else max(12, min(38, len(nome) + 16))
        ws.column_dimensions[get_column_letter(i)].width = largura

    ws.freeze_panes = "A2"
    if linhas:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}{len(linhas) + 1}"
    return ws


def main():
    raiz = Path(sys.argv[1] if len(sys.argv) > 1 else ".")
    saida = Path(sys.argv[2] if len(sys.argv) > 2 else "banco-criativo.xlsx")

    criativos = carregar(raiz / "banco" / "criativos.json")
    virais = carregar(raiz / "banco" / "virais.json")
    publico = carregar(raiz / "banco" / "publico.json")
    angulos = carregar(raiz / "mineracao" / "angulos.json")
    clickbait = carregar(raiz / "mineracao" / "clickbait.json")
    formatos = carregar(raiz / "mineracao" / "formatos.json")
    benchmark = carregar(raiz / "mineracao" / "benchmark.json")
    fatias_banco = carregar(raiz / "banco-de-fatias.json")

    if isinstance(benchmark, dict):
        anunciantes = benchmark.get("anunciantes", [])
        personas_bm = benchmark.get("micro_personas", [])
        fatias_bm = benchmark.get("fatias", [])
    else:
        anunciantes, personas_bm, fatias_bm = [], [], []

    fatias = fatias_bm + (fatias_banco if isinstance(fatias_banco, list) else [])

    wb = Workbook()
    wb.remove(wb.active)

    # --- Criativos: 1 linha por ad dissecado -------------------------------
    cols = ["id", "anunciante", "oferta", "dias_no_ar", "anuncios_ativos",
            "angulo", "avatar", "formato", "hook_visual", "hook_copy",
            "micro_persona_id", "micro_persona", "nivel_consciencia",
            "estado_cognitivo", "emocao", "crenca_respeitada", "transcricao_origem",
            "linha_vermelha", "link_anuncio", "link_biblioteca", "link_funil"]
    linhas = [[
        pegar(c, "id"), pegar(c, "fonte", "anunciante"), pegar(c, "oferta"),
        pegar(c, "fonte", "dias_no_ar"), pegar(c, "fonte", "anuncios_ativos"),
        pegar(c, "angulo", "descricao"), pegar(c, "avatar", "descricao"),
        pegar(c, "formato", "macro"), pegar(c, "hook", "visual"),
        pegar(c, "hook", "copy"), pegar(c, "micro_persona_id"),
        pegar(c, "micro_persona"), pegar(c, "nivel_consciencia"),
        pegar(c, "estado_cognitivo"), pegar(c, "emocao"),
        pegar(c, "crenca_respeitada"), pegar(c, "transcricao", "origem"),
        pegar(c, "linha_vermelha"), pegar(c, "fonte", "link_anuncio"),
        pegar(c, "fonte", "link_biblioteca"), pegar(c, "fonte", "link_funil"),
    ] for c in criativos]
    escrever_aba(wb, "Criativos", cols, linhas,
                 ("link_anuncio", "link_biblioteca", "link_funil"))

    # --- Angulos -----------------------------------------------------------
    cols = ["id", "descricao", "mecanismo", "principios", "micro_persona_id",
            "micro_persona", "crenca_respeitada", "nivel_consciencia",
            "estado_cognitivo", "repeticao_pct", "repeticao_pct_controle",
            "indice_discriminacao", "amostra_n", "adaptabilidade",
            "variacao_negativa", "linha_vermelha", "status",
            "link_anuncio", "link_biblioteca", "link_funil"]
    linhas = [[
        pegar(a, "id"), pegar(a, "descricao"), pegar(a, "mecanismo"),
        pegar(a, "principios"), pegar(a, "micro_persona_id"),
        pegar(a, "micro_persona"), pegar(a, "crenca_respeitada"),
        pegar(a, "nivel_consciencia"), pegar(a, "estado_cognitivo"),
        pegar(a, "repeticao_pct"), pegar(a, "repeticao_pct_controle"),
        pegar(a, "indice_discriminacao"), pegar(a, "amostra_n"),
        pegar(a, "adaptabilidade"), pegar(a, "variacao_negativa"),
        pegar(a, "linha_vermelha"), pegar(a, "status"),
        pegar(a, "fonte", "link_anuncio"), pegar(a, "fonte", "link_biblioteca"),
        pegar(a, "fonte", "link_funil"),
    ] for a in angulos]
    escrever_aba(wb, "Angulos", cols, linhas,
                 ("link_anuncio", "link_biblioteca", "link_funil"))

    # --- Avatares: extraidos dos criativos dissecados ----------------------
    cols = ["avatar_id", "descricao", "genero", "faixa_etaria", "papel",
            "voz", "tratamento", "ocorrencias", "micro_persona_id",
            "angulos_em_que_aparece", "exemplo_link_anuncio"]
    vistos = {}
    for c in criativos:
        chave = pegar(c, "avatar", "descricao") or "(nao informado)"
        registro = vistos.setdefault(chave, {
            "id": pegar(c, "avatar", "id") or f"AVT-{len(vistos) + 1:03d}",
            "genero": pegar(c, "avatar", "genero"),
            "faixa": pegar(c, "avatar", "faixa_etaria"),
            "papel": pegar(c, "avatar", "papel"),
            "voz": pegar(c, "avatar", "voz"),
            "trat": pegar(c, "avatar", "tratamento"),
            "n": 0, "mp": pegar(c, "micro_persona_id"),
            "ang": set(), "link": pegar(c, "fonte", "link_anuncio"),
        })
        registro["n"] += 1
        if pegar(c, "angulo", "id"):
            registro["ang"].add(pegar(c, "angulo", "id"))
    linhas = [[
        v["id"], k, v["genero"], v["faixa"], v["papel"], v["voz"], v["trat"],
        v["n"], v["mp"], ", ".join(sorted(v["ang"])), v["link"],
    ] for k, v in sorted(vistos.items(), key=lambda x: -x[1]["n"])]
    escrever_aba(wb, "Avatares", cols, linhas, ("exemplo_link_anuncio",))

    # --- Formatos ----------------------------------------------------------
    cols = ["id", "descricao", "mecanismo", "micro_persona_id", "micro_persona",
            "repeticao_pct", "repeticao_pct_controle", "indice_discriminacao",
            "amostra_n", "sinal_organico", "custo_producao", "adaptabilidade",
            "hook_0a3s", "linha_vermelha", "status",
            "link_anuncio", "link_biblioteca", "link_funil"]
    linhas = [[
        pegar(f, "id"), pegar(f, "descricao"), pegar(f, "mecanismo"),
        pegar(f, "micro_persona_id"), pegar(f, "micro_persona"),
        pegar(f, "repeticao_pct"), pegar(f, "repeticao_pct_controle"),
        pegar(f, "indice_discriminacao"), pegar(f, "amostra_n"),
        pegar(f, "sinal_organico"), pegar(f, "custo_producao"),
        pegar(f, "adaptabilidade"), pegar(f, "hook_0a3s"),
        pegar(f, "linha_vermelha"), pegar(f, "status"),
        pegar(f, "fonte", "link_anuncio"), pegar(f, "fonte", "link_biblioteca"),
        pegar(f, "fonte", "link_funil"),
    ] for f in formatos]
    escrever_aba(wb, "Formatos", cols, linhas,
                 ("link_anuncio", "link_biblioteca", "link_funil"))

    # --- Hooks (clickbait / hook visual) -----------------------------------
    cols = ["id", "descricao", "mecanismo", "principios", "tipo_gancho",
            "micro_persona_id", "micro_persona", "captura", "cenario", "elenco",
            "props", "enquadramento", "tratamento", "custo", "tempo_producao",
            "adaptacao_nicho", "repeticao_pct", "indice_discriminacao",
            "linha_vermelha", "status", "link_anuncio", "link_biblioteca",
            "link_funil"]
    linhas = [[
        pegar(h, "id"), pegar(h, "descricao"), pegar(h, "mecanismo"),
        pegar(h, "principios"), pegar(h, "tipo_gancho"),
        pegar(h, "micro_persona_id"), pegar(h, "micro_persona"),
        pegar(h, "captura"), pegar(h, "receita_replicacao", "cenario"),
        pegar(h, "receita_replicacao", "elenco"),
        pegar(h, "receita_replicacao", "props"),
        pegar(h, "receita_replicacao", "enquadramento"),
        pegar(h, "receita_replicacao", "tratamento"),
        pegar(h, "receita_replicacao", "custo"),
        pegar(h, "receita_replicacao", "tempo_producao"),
        pegar(h, "receita_replicacao", "adaptacao_nicho"),
        pegar(h, "repeticao_pct"), pegar(h, "indice_discriminacao"),
        pegar(h, "linha_vermelha"), pegar(h, "status"),
        pegar(h, "fonte", "link_anuncio"), pegar(h, "fonte", "link_biblioteca"),
        pegar(h, "fonte", "link_funil"),
    ] for h in clickbait]
    escrever_aba(wb, "Hooks", cols, linhas,
                 ("link_anuncio", "link_biblioteca", "link_funil"))

    # --- Publico: micro-personas do estudo + do benchmark ------------------
    cols = ["id", "nome", "papel_ameacado", "quem_compra", "voz", "faixa_etaria",
            "gatilho_emocional", "dor_central", "desejo_central",
            "objecao_principal", "crencas", "linguagem_literal", "onde_vive",
            "ocorrencias", "pct_grupo_a", "pct_grupo_b", "indice_discriminacao",
            "confianca", "fonte_evidencia"]
    fonte_publico = (publico if isinstance(publico, list) else []) + personas_bm
    linhas = [[
        pegar(p, "id"), pegar(p, "nome"), pegar(p, "papel_ameacado"),
        pegar(p, "quem_compra"), pegar(p, "voz"), pegar(p, "faixa_etaria"),
        pegar(p, "gatilho_emocional"), pegar(p, "dor_central"),
        pegar(p, "desejo_central"), pegar(p, "objecao_principal"),
        pegar(p, "crencas"), pegar(p, "linguagem_literal"), pegar(p, "onde_vive"),
        pegar(p, "ocorrencias"), pegar(p, "pct_grupo_a"), pegar(p, "pct_grupo_b"),
        pegar(p, "indice_discriminacao"), pegar(p, "confianca"),
        pegar(p, "fonte_evidencia"),
    ] for p in fonte_publico]
    escrever_aba(wb, "Publico", cols, linhas)

    # --- Virais organicos ---------------------------------------------------
    cols = ["id", "plataforma", "titulo", "views", "data", "formato_macro",
            "angulo", "avatar", "hook_0a3s", "gancho_copy", "mecanismo",
            "sinal_organico", "transferivel_para_ads", "custo_producao",
            "micro_persona_id", "linha_vermelha", "link"]
    linhas = [[
        pegar(v, "id"), pegar(v, "plataforma"), pegar(v, "titulo"),
        pegar(v, "views"), pegar(v, "data"), pegar(v, "formato_macro"),
        pegar(v, "angulo"), pegar(v, "avatar"), pegar(v, "hook_0a3s"),
        pegar(v, "gancho_copy"), pegar(v, "mecanismo"), pegar(v, "sinal_organico"),
        pegar(v, "transferivel_para_ads"), pegar(v, "custo_producao"),
        pegar(v, "micro_persona_id"), pegar(v, "linha_vermelha"),
        pegar(v, "link"),
    ] for v in virais]
    escrever_aba(wb, "Virais", cols, linhas, ("link",))

    # --- Anunciantes --------------------------------------------------------
    cols = ["nome", "classificacao", "dias_no_ar_max", "criativos_unicos",
            "linhas_na_biblioteca", "pais", "ticket", "escada_precos",
            "mecanismo", "angulos_usados", "formatos_usados",
            "micro_personas_atacadas", "nivel_consciencia",
            "link_biblioteca", "link_funil"]
    linhas = [[
        pegar(a, "nome"), pegar(a, "classificacao"), pegar(a, "dias_no_ar_max"),
        pegar(a, "criativos_unicos_estimados"), pegar(a, "linhas_na_biblioteca"),
        pegar(a, "pais"), pegar(a, "ticket"), pegar(a, "escada_precos"),
        pegar(a, "mecanismo"), pegar(a, "angulos_usados"),
        pegar(a, "formatos_usados"), pegar(a, "micro_personas_atacadas"),
        pegar(a, "nivel_consciencia"), pegar(a, "link_biblioteca"),
        pegar(a, "link_funil"),
    ] for a in anunciantes]
    escrever_aba(wb, "Anunciantes", cols, linhas, ("link_biblioteca", "link_funil"))

    # --- Fatias -------------------------------------------------------------
    cols = ["id", "origem", "micro_persona_id", "micro_persona", "angulo",
            "formato", "densidade", "anunciantes_na_fatia", "descricao",
            "custo_producao", "prioridade", "status"]
    linhas = [[
        pegar(f, "id"), pegar(f, "origem"), pegar(f, "micro_persona_id"),
        pegar(f, "micro_persona"), pegar(f, "angulo"), pegar(f, "formato"),
        pegar(f, "densidade"), pegar(f, "anunciantes_na_fatia"),
        pegar(f, "descricao"), pegar(f, "custo_producao"),
        pegar(f, "prioridade"), pegar(f, "status"),
    ] for f in fatias]
    escrever_aba(wb, "Fatias", cols, linhas)

    # --- Resumo (formulas, recalculam sozinhas) ----------------------------
    ws = wb.create_sheet("Resumo", 0)
    ws["A1"] = "Banco Criativo — Resumo"
    ws["A1"].font = Font(name=FONTE, bold=True, size=14)
    ws["A3"] = "Aba"
    ws["B3"] = "Itens"
    ws["C3"] = "Linha vermelha"
    for celula in ws["A3:C3"][0]:
        celula.fill = CAB_FILL
        celula.font = CAB_FONT

    mapa = [("Criativos", "R"), ("Angulos", "P"), ("Formatos", "N"),
            ("Hooks", "S"), ("Publico", None), ("Virais", "P"),
            ("Anunciantes", None), ("Fatias", None)]
    linha = 4
    for aba, col_lv in mapa:
        ws[f"A{linha}"] = aba
        ws[f"B{linha}"] = f"=COUNTA({aba}!A2:A5000)"
        ws[f"C{linha}"] = (f'=COUNTIF({aba}!{col_lv}2:{col_lv}5000,TRUE)'
                           if col_lv else "n/a")
        for col in "ABC":
            ws[f"{col}{linha}"].font = CEL_FONT
        linha += 1

    ws[f"A{linha + 1}"] = ("Contagens sao formulas: crescem sozinhas quando voce "
                           "reprocessar os JSONs e regerar a planilha.")
    ws[f"A{linha + 1}"].font = Font(name=FONTE, size=9, italic=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18

    wb.save(saida)

    print(f"planilha gerada: {saida}")
    print(f"  Criativos   {len(criativos):4d}")
    print(f"  Angulos     {len(angulos):4d}")
    print(f"  Avatares    {len(vistos):4d}")
    print(f"  Formatos    {len(formatos):4d}")
    print(f"  Hooks       {len(clickbait):4d}")
    print(f"  Publico     {len(fonte_publico):4d}")
    print(f"  Virais      {len(virais):4d}")
    print(f"  Anunciantes {len(anunciantes):4d}")
    print(f"  Fatias      {len(fatias):4d}")
    vazias = [n for n, q in [("criativos", criativos), ("angulos", angulos),
                             ("formatos", formatos), ("hooks", clickbait),
                             ("publico", fonte_publico), ("virais", virais)] if not q]
    if vazias:
        print(f"  abas sem dado (fonte ausente): {', '.join(vazias)}")


if __name__ == "__main__":
    main()
